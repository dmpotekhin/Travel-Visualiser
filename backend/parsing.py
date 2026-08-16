"""Parse uploaded CSV / Excel route tables.

Expected columns (matched by header keyword, case-insensitive):
  - route:   "Маршрут (кратко)" / "Маршрут" / "Route"
  - year:    "Год" / "Year"
  - note:    "Примечания" / "Note"
  - distance:"Общее расстояние, км" / "Distance" (optional, user's own figure)
"""
from __future__ import annotations

import csv
import io
from typing import Optional

from openpyxl import load_workbook

_HEADER_KEYWORDS = {
    "route": ("маршрут", "route", "путь", "маршру"),
    "year": ("год", "year"),
    "note": ("примечан", "note", "коммент", "comment"),
    "distance": ("расстоян", "distance", "км"),
}


def _find_columns(headers: list[str]) -> dict[str, Optional[int]]:
    cols = {"route": None, "year": None, "note": None, "distance": None}
    for i, h in enumerate(headers):
        hl = str(h).strip().lower()
        for key, kws in _HEADER_KEYWORDS.items():
            if cols[key] is None and any(k in hl for k in kws):
                cols[key] = i
    return cols


def _cell(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx]).strip()


def _to_int(value: str) -> Optional[int]:
    try:
        return int(float(value.replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _to_float(value: str) -> Optional[float]:
    try:
        return float(value.replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _rows_to_routes(rows: list[list], start_idx: int) -> list[dict]:
    headers = rows[start_idx]
    cols = _find_columns(headers)
    if cols["route"] is None:
        raise ValueError("Не найдена колонка с маршрутом (ищите заголовок «Маршрут»).")

    routes = []
    for row in rows[start_idx + 1 :]:
        route = _cell(row, cols["route"])
        if not route:
            continue
        routes.append(
            {
                "route": route,
                "year": _to_int(_cell(row, cols["year"])),
                "note": _cell(row, cols["note"]) or None,
                "declared_km": _to_float(_cell(row, cols["distance"])),
            }
        )
    if not routes:
        raise ValueError("В файле не найдено ни одного маршрута.")
    return routes


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [r for r in reader]
    # drop fully empty rows
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if len(rows) < 2:
        raise ValueError("CSV-файл пуст или не содержит заголовка.")
    return _rows_to_routes(rows, 0)


def _parse_excel(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    # drop fully empty rows
    rows = [r for r in rows if any((c or "") != "" for c in r)]
    if len(rows) < 2:
        raise ValueError("Excel-файл пуст или не содержит заголовка.")
    return _rows_to_routes(rows, 0)


def parse_upload(filename: str, content: bytes) -> list[dict]:
    """Dispatch by extension and return a list of route dicts."""
    name = filename.lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_excel(content)
    if name.endswith(".xls"):
        raise ValueError("Формат .xls не поддерживается — сохраните файл как .xlsx или .csv.")
    raise ValueError("Поддерживаются только файлы .csv и .xlsx.")
